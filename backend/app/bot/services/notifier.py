import logging
from typing import Any, List, Optional
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from aiogram import Bot
from aiogram.types import BufferedInputFile

from app.bot.config import bot_settings
from app.bot.keyboards.warehouse import get_order_warehouse_keyboard
from app.models.order import OrderStatus

logger = logging.getLogger(__name__)


def _format_qty(qty: Optional[float]) -> str:
    """Format quantity cleanly removing redundant trailing decimal zeros."""
    if qty is None:
        return "0"
    val = float(qty)
    if val.is_integer():
        return str(int(val))
    return f"{val:.2f}".rstrip("0").rstrip(".")

def generate_excel_order(order_id: int, bar_name: str, items: List[Any]) -> BufferedInputFile:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заказ {order_id}"
    
    clean_name = bar_name.strip()
    if clean_name.startswith("Кофейня"):
        clean_name = clean_name[len("Кофейня"):].strip()
    clean_name = clean_name.strip("«»\"' ")
    if not clean_name:
        clean_name = bar_name

    cyan_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    peach_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    now = datetime.now()
    current_date = f"{now.day}.{now.month}"
    
    ws.append(["Дата отправки", current_date, "", ""])
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.fill = cyan_fill
        cell.border = thin_border
        cell.alignment = center_align
        cell.font = Font(size=12)

    ws.append(["Кофе бар", clean_name, "Склад", "Бар"])
    for col in range(1, 5):
        cell = ws.cell(row=2, column=col)
        cell.fill = peach_fill
        cell.border = thin_border
        cell.alignment = center_align
        cell.font = Font(size=12)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    row_idx = 3
    for item in items:
        if isinstance(item, dict):
            name = item.get("name") or item.get("product_name") or "Товар"
            qty = item.get("confirmed_qty")
            if qty is None:
                qty = item.get("requested_qty", 0)
            unit = item.get("unit") or ""
        else:
            name = getattr(item, "name", None)
            if not name and hasattr(item, "product") and item.product:
                name = item.product.name
            name = name or "Товар"
            qty = getattr(item, "confirmed_qty", None)
            if qty is None:
                qty = getattr(item, "requested_qty", 0)
            unit = getattr(item, "unit", None)
            if not unit and hasattr(item, "product") and item.product:
                unit = item.product.unit
            unit = unit or ""

        if qty is not None and float(qty) > 0:
            qty_str = _format_qty(qty)
            if unit:
                qty_str += f" {unit}"
            ws.append([name, qty_str, "", ""])
            for col in range(1, 5):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = center_align
                cell.font = Font(size=12)
            row_idx += 1
            
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Заказ_{order_id}_{clean_name}.xlsx".replace(" ", "_")
    return BufferedInputFile(out.read(), filename=filename)



def format_order_message(
    order_id: int,
    bar_name: str,
    items: List[Any],
    out_of_stock_items: Optional[List[Any]] = None,
    status: OrderStatus = OrderStatus.PENDING,
    packer_name: Optional[str] = None,
    shipped_by: Optional[str] = None,
    include_items: bool = False,
) -> str:
    """Format warehouse notification message according to specification.

    Example:
    📦 Новый заказ #1042 — Кофейня «Центр»
    ──────────────────────────────
    • Стакан 0,3 — 50 шт.
    • Сироп «Солёная карамель» — 2 бут.
    • Молоко кокосовое — 6 шт.
    • Салфетки простые — 4 уп.
    ──────────────────────────────
    ⚠️ В стопе (не вошло):
    • Сахар в стиках — 0 шт.
    """
    # Clean and normalize bar name (e.g. 'Кофейня «Центр»' or 'Центр' -> 'Центр')
    clean_name = bar_name.strip()
    if clean_name.startswith("Кофейня"):
        clean_name = clean_name[len("Кофейня"):].strip()
    clean_name = clean_name.strip("«»\"' ")
    if not clean_name:
        clean_name = bar_name

    if status == OrderStatus.SHIPPED:
        lines = [
            f"✅ Заказ #{order_id} отгружен",
            f"Кофейня: «{clean_name}»",
        ]
        if shipped_by:
            lines.append(f"Отгрузил: {shipped_by}")
        lines.append("──────────────────────────────")
    elif status == OrderStatus.CANCELLED:
        lines = [
            f"❌ Заказ #{order_id} отменен",
            f"Кофейня: «{clean_name}»",
            "──────────────────────────────",
        ]
    else:
        lines = [
            f"📦 Новый заказ #{order_id} — Кофейня «{clean_name}»",
            "──────────────────────────────",
        ]

    # Confirmed/active items list
    if include_items:
        confirmed_lines: List[str] = []
        for item in items:
            # Support both object attributes and dict keys
            if isinstance(item, dict):
                name = item.get("name") or item.get("product_name") or "Товар"
                qty = item.get("confirmed_qty")
                if qty is None:
                    qty = item.get("requested_qty", 0)
                unit = item.get("unit") or "шт."
            else:
                name = getattr(item, "name", None)
                if not name and hasattr(item, "product") and item.product:
                    name = item.product.name
                name = name or "Товар"
                qty = getattr(item, "confirmed_qty", None)
                if qty is None:
                    qty = getattr(item, "requested_qty", 0)
                unit = getattr(item, "unit", None)
                if not unit and hasattr(item, "product") and item.product:
                    unit = item.product.unit
                unit = unit or "шт."
    
            # Only list in main section if confirmed_qty > 0 or not marked as 0
            if qty is not None and float(qty) > 0:
                confirmed_lines.append(f"• {name} — {_format_qty(qty)} {unit}")
    
        if confirmed_lines:
            lines.extend(confirmed_lines)
        else:
            lines.append("• (Нет подтвержденных позиций)")

    # Out-of-stock items list
    oos_lines: List[str] = []
    if out_of_stock_items:
        for oos in out_of_stock_items:
            if isinstance(oos, dict):
                name = oos.get("name") or "Товар"
                unit = oos.get("unit") or "шт."
            else:
                name = getattr(oos, "name", None)
                if not name and hasattr(oos, "product") and oos.product:
                    name = oos.product.name
                name = name or "Товар"
                unit = getattr(oos, "unit", None)
                if not unit and hasattr(oos, "product") and oos.product:
                    unit = oos.product.unit
                unit = unit or "шт."
            oos_lines.append(f"• {name} — 0 {unit}")

    # Also check items in `items` that had confirmed_qty == 0 and were not already in oos_lines
    for item in items:
        if isinstance(item, dict):
            qty = item.get("confirmed_qty")
            name = item.get("name") or "Товар"
            unit = item.get("unit") or "шт."
        else:
            qty = getattr(item, "confirmed_qty", None)
            name = getattr(item, "name", None)
            if not name and hasattr(item, "product") and item.product:
                name = item.product.name
            name = name or "Товар"
            unit = getattr(item, "unit", None)
            if not unit and hasattr(item, "product") and item.product:
                unit = item.product.unit
            unit = unit or "шт."

        if qty is not None and float(qty) == 0:
            line = f"• {name} — 0 {unit}"
            if line not in oos_lines:
                oos_lines.append(line)

    if oos_lines:
        lines.append("──────────────────────────────")
        lines.append("⚠️ В стопе (не вошло): ")
        lines.extend(oos_lines)

    # If packing info is present
    if status == OrderStatus.PACKING and packer_name:
        lines.append("──────────────────────────────")
        lines.append(f"👨‍🍳 В сборке: {packer_name}")

    return "\n".join(lines)


async def send_order_to_warehouse(
    order_id: int,
    bar_name: str,
    items: List[Any],
    out_of_stock_items: Optional[List[Any]] = None,
    bot: Optional[Bot] = None,
    chat_id: Optional[int] = None,
) -> Optional[Any]:
    """Send order notification to warehouse telegram chat.

    Gracefully logs and ignores errors if bot token or chat ID is not set.
    """
    token = bot_settings.TOKEN
    target_chat_id = chat_id or bot_settings.WAREHOUSE_CHAT_ID

    if not token or not target_chat_id:
        logger.info(
            "Telegram bot token or warehouse chat ID is not configured. Skipping warehouse notification."
        )
        return None

    close_bot = False
    if bot is None:
        bot = Bot(token=token)
        close_bot = True

    try:
        text = format_order_message(
            order_id=order_id,
            bar_name=bar_name,
            items=items,
            out_of_stock_items=out_of_stock_items,
            status=OrderStatus.PENDING,
            include_items=False,
        )
        keyboard = get_order_warehouse_keyboard(order_id, status=OrderStatus.PENDING)

        document = generate_excel_order(order_id, bar_name, items)
        
        msg = await bot.send_document(
            chat_id=target_chat_id,
            document=document,
            caption=text,
            reply_markup=keyboard,
        )
        return msg
    except Exception as e:
        logger.error(f"Failed to send order notification to warehouse chat: {e}", exc_info=True)
        return None
    finally:
        if close_bot:
            await bot.session.close()
